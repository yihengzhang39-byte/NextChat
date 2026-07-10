#!/usr/bin/env python3
"""Batch Excel evaluator for Iflytek image-only prompts.

The script reads image paths from Excel, remaps them to a local image directory
by file name/stem, and calls the existing local NextChat backend route. Iflytek
WebSocket signing, payload conversion, SSE conversion, and response parsing
remain centralized in the app.
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import json
import mimetypes
import os
import re
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urljoin


DEFAULT_BASE_URL = "http://127.0.0.1:3000"
DEFAULT_API_PATH = "/api/iflytek/v1/chat/completions"
DEFAULT_MODEL = "image@Iflytek"
DEFAULT_IMAGE_ROOT = Path(r"D:\test_datamodel\图生文\图生文\图片")
DEFAULT_IMAGE_COLUMN = "C"
DEFAULT_REPLY_COLUMN = "D"
DEFAULT_STATUS_COLUMN = "E"
DEFAULT_LATENCY_COLUMN = "F"
DEFAULT_ERROR_COLUMN = "G"
DEFAULT_SID_COLUMN = "H"
DEFAULT_START_ROW = 2
DEFAULT_CONCURRENCY = 3
DEFAULT_TIMEOUT_SECONDS = 180.0
DEFAULT_RETRIES = 2
DEFAULT_PROMPT = "请根据图片内容生成对应的文字描述。"
SUPPORTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
EXTENSION_PRIORITY = [".jpg", ".jpeg", ".png", ".webp"]


@dataclass(frozen=True)
class Job:
    row: int
    image_path: Path


@dataclass(frozen=True)
class Result:
    row: int
    status: str
    reply: str = ""
    latency_seconds: float | None = None
    error: str = ""
    sid: str = ""


@dataclass(frozen=True)
class ColumnLayout:
    status: str
    latency: str
    error: str
    sid: str


@dataclass(frozen=True)
class CollectionStats:
    total_rows: int
    valid_samples: int
    image_missing_count: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read image paths from an Excel column, remap them into a local "
            "image directory, and write independent Iflytek image answers."
        )
    )
    parser.add_argument("--input", required=True, help="Input .xlsx path.")
    parser.add_argument(
        "--output",
        help=(
            "Output .xlsx path. Defaults to '<input stem>_with_replies.xlsx'. "
            "The input file is never overwritten."
        ),
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to the active worksheet.",
    )
    parser.add_argument(
        "--image-root",
        default=str(DEFAULT_IMAGE_ROOT),
        help=f"Directory containing the real image files. Default: {DEFAULT_IMAGE_ROOT}.",
    )
    parser.add_argument(
        "--image-column",
        default=DEFAULT_IMAGE_COLUMN,
        help=f"Image path column letter. Default: {DEFAULT_IMAGE_COLUMN}.",
    )
    parser.add_argument(
        "--reply-column",
        default=DEFAULT_REPLY_COLUMN,
        help=f"Reply column letter. Default: {DEFAULT_REPLY_COLUMN}.",
    )
    parser.add_argument(
        "--prompt",
        default=DEFAULT_PROMPT,
        help=(
            "Instruction sent with every image. Use an empty string to send only "
            f"the image. Default: {DEFAULT_PROMPT}"
        ),
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=DEFAULT_CONCURRENCY,
        help=f"Maximum concurrent requests. Default: {DEFAULT_CONCURRENCY}.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_TIMEOUT_SECONDS,
        help=f"Per-row timeout in seconds. Default: {DEFAULT_TIMEOUT_SECONDS:g}.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_RETRIES,
        help=f"Retries after the first failed attempt. Default: {DEFAULT_RETRIES}.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-request rows even when the reply column already has text.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=DEFAULT_START_ROW,
        help=f"First data row. Default: {DEFAULT_START_ROW}.",
    )
    parser.add_argument(
        "--base-url",
        default=os.getenv("NEXTCHAT_BASE_URL", DEFAULT_BASE_URL),
        help=f"Local NextChat base URL. Default: {DEFAULT_BASE_URL}.",
    )
    parser.add_argument(
        "--api-path",
        default=DEFAULT_API_PATH,
        help=f"Iflytek backend API path. Default: {DEFAULT_API_PATH}.",
    )
    parser.add_argument(
        "--model",
        default=os.getenv("BATCH_IFLYTEK_MODEL", DEFAULT_MODEL),
        help=f"Model sent to the local backend. Default: {DEFAULT_MODEL}.",
    )
    parser.add_argument(
        "--access-code",
        default=os.getenv("NEXTCHAT_ACCESS_CODE", ""),
        help="Optional NextChat access code when CODE is enabled.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not call the backend; write deterministic fake replies.",
    )
    return parser.parse_args()


def normalize_column(value: str, label: str) -> str:
    column = value.strip().upper()
    if not column.isalpha():
        raise ValueError(f"{label} must be an Excel column letter.")
    return column


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_with_replies{input_path.suffix}")


def resolve_paths(args: argparse.Namespace) -> tuple[Path, Path, Path]:
    input_path = Path(args.input).expanduser().resolve()
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else default_output_path(input_path).resolve()
    )
    image_root = Path(args.image_root).expanduser().resolve()

    if input_path == output_path:
        raise ValueError("Output path must be different from input path.")

    return input_path, output_path, image_root


def load_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install it with "
            "'python -m pip install openpyxl'."
        ) from exc

    return load_workbook(path)


def get_sheet(workbook: Any, sheet_name: str | None):
    if not sheet_name:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Worksheet '{sheet_name}' not found. Available: {available}")
    return workbook[sheet_name]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_path_text(value: Any) -> str:
    text = cell_text(value)
    return text.strip().strip('"').strip("'").strip("\u201c\u201d\u2018\u2019").strip()


def candidate_image_paths(raw_value: Any, image_root: Path) -> list[Path]:
    cleaned = clean_path_text(raw_value)
    if not cleaned:
        return []

    original = Path(cleaned).expanduser()
    original_name = original.name
    original_stem = original.stem
    candidates: list[Path] = []

    def add_candidate(path: Path) -> None:
        resolved = path.resolve(strict=False)
        if resolved not in candidates:
            candidates.append(resolved)

    # Prefer the configured real image directory. This supports old absolute
    # paths whose directory or extension no longer matches the actual files.
    if original_name:
        add_candidate(image_root / original_name)
    if original_stem:
        suffixes = [original.suffix.lower(), *EXTENSION_PRIORITY]
        for suffix in suffixes:
            if suffix in SUPPORTED_IMAGE_EXTENSIONS:
                add_candidate(image_root / f"{original_stem}{suffix}")

    if not original.is_absolute():
        add_candidate(image_root / original)
    add_candidate(original)

    return candidates


def resolve_image_path(raw_value: Any, image_root: Path) -> Path | None:
    for candidate in candidate_image_paths(raw_value, image_root):
        if candidate.exists() and candidate.is_file():
            return candidate
    candidates = candidate_image_paths(raw_value, image_root)
    return candidates[0] if candidates else None


def validate_image_path(path: Path | None) -> str:
    if path is None:
        return "C列图片路径为空。"
    if not path.exists():
        return f"图片文件不存在：{path}"
    if not path.is_file():
        return f"图片路径不是文件：{path}"
    if path.suffix.lower() not in SUPPORTED_IMAGE_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_IMAGE_EXTENSIONS))
        return f"图片格式不支持：{path}；支持格式：{supported}"
    return ""


def is_column_empty(sheet: Any, column: str) -> bool:
    for row in range(1, sheet.max_row + 1):
        if cell_text(sheet[f"{column}{row}"].value):
            return False
    return True


def can_reuse_status_columns(sheet: Any, reply_column: str) -> bool:
    expected = {
        DEFAULT_STATUS_COLUMN: "status",
        DEFAULT_LATENCY_COLUMN: "latency_seconds",
        DEFAULT_ERROR_COLUMN: "error",
        DEFAULT_SID_COLUMN: "sid",
    }
    for column, header in expected.items():
        if column == reply_column:
            return False
        current_header = cell_text(sheet[f"{column}1"].value)
        if current_header == header:
            continue
        if is_column_empty(sheet, column):
            continue
        return False
    return True


def choose_status_columns(sheet: Any, reply_column: str) -> ColumnLayout:
    if can_reuse_status_columns(sheet, reply_column):
        return ColumnLayout(
            status=DEFAULT_STATUS_COLUMN,
            latency=DEFAULT_LATENCY_COLUMN,
            error=DEFAULT_ERROR_COLUMN,
            sid=DEFAULT_SID_COLUMN,
        )

    from openpyxl.utils import column_index_from_string, get_column_letter

    start = max(sheet.max_column, column_index_from_string(reply_column)) + 1
    return ColumnLayout(
        status=get_column_letter(start),
        latency=get_column_letter(start + 1),
        error=get_column_letter(start + 2),
        sid=get_column_letter(start + 3),
    )


def collect_jobs(
    sheet: Any,
    start_row: int,
    image_column: str,
    reply_column: str,
    image_root: Path,
    force: bool,
) -> tuple[list[Job], list[Result], dict[int, Result], CollectionStats]:
    jobs: list[Job] = []
    skipped: list[Result] = []
    invalid: dict[int, Result] = {}
    total_rows = max(0, sheet.max_row - start_row + 1)
    valid_samples = 0
    image_missing_count = 0

    for row in range(start_row, sheet.max_row + 1):
        image_path = resolve_image_path(sheet[f"{image_column}{row}"].value, image_root)
        existing_reply = cell_text(sheet[f"{reply_column}{row}"].value)

        if existing_reply and not force:
            skipped.append(Result(row=row, status="skipped", reply=existing_reply))
            if image_path is not None:
                valid_samples += 1
            continue

        image_error = validate_image_path(image_path)
        if image_error:
            image_missing_count += 1
            invalid[row] = Result(row=row, status="failed", error=image_error)
            continue

        assert image_path is not None
        valid_samples += 1
        jobs.append(Job(row=row, image_path=image_path))

    stats = CollectionStats(
        total_rows=total_rows,
        valid_samples=valid_samples,
        image_missing_count=image_missing_count,
    )
    return jobs, skipped, invalid, stats


async def run_jobs(
    jobs: list[Job],
    args: argparse.Namespace,
) -> dict[int, Result]:
    semaphore = asyncio.Semaphore(args.concurrency)
    results: dict[int, Result] = {}
    api_url = build_api_url(args.base_url, args.api_path)

    async def run_one(job: Job) -> None:
        async with semaphore:
            if args.dry_run:
                started_at = time.perf_counter()
                await asyncio.sleep(0)
                results[job.row] = Result(
                    row=job.row,
                    status="success",
                    reply=f"[dry-run] {job.image_path.name}",
                    latency_seconds=time.perf_counter() - started_at,
                )
                return

            results[job.row] = await request_with_retries(
                job=job,
                api_url=api_url,
                model=args.model,
                prompt=args.prompt,
                access_code=args.access_code,
                timeout_seconds=args.timeout,
                retries=args.retries,
            )

    await asyncio.gather(*(run_one(job) for job in jobs))
    return results


async def request_with_retries(
    job: Job,
    api_url: str,
    model: str,
    prompt: str,
    access_code: str,
    timeout_seconds: float,
    retries: int,
) -> Result:
    attempts = max(0, retries) + 1
    last_error = ""
    last_sid = ""
    started_at = time.perf_counter()

    for attempt in range(1, attempts + 1):
        try:
            attempt_started_at = time.perf_counter()
            reply = await asyncio.wait_for(
                asyncio.to_thread(
                    call_nextchat_iflytek,
                    api_url,
                    model,
                    job.image_path,
                    prompt,
                    access_code,
                    timeout_seconds,
                ),
                timeout=timeout_seconds,
            )
            return Result(
                row=job.row,
                status="success",
                reply=reply,
                latency_seconds=time.perf_counter() - attempt_started_at,
            )
        except Exception as exc:  # noqa: BLE001 - keep one-row failures isolated.
            last_error = format_row_error(exc)
            if attempt < attempts:
                await asyncio.sleep(min(2 ** (attempt - 1), 5))

    return Result(
        row=job.row,
        status="failed",
        latency_seconds=time.perf_counter() - started_at,
        error=last_error,
        sid=extract_iflytek_sid(last_error),
    )


def build_api_url(base_url: str, api_path: str) -> str:
    base = base_url.rstrip("/") + "/"
    path = api_path.lstrip("/")
    return urljoin(base, path)


def image_to_data_url(path: Path) -> str:
    image_error = validate_image_path(path)
    if image_error:
        raise RuntimeError(image_error)

    mime_type = mimetypes.guess_type(path.name)[0] or extension_mime_type(path)
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def extension_mime_type(path: Path) -> str:
    return {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
    }.get(path.suffix.lower(), "application/octet-stream")


def call_nextchat_iflytek(
    api_url: str,
    model: str,
    image_path: Path,
    prompt: str,
    access_code: str,
    timeout_seconds: float,
) -> str:
    content: list[dict[str, Any]] = [
        {
            "type": "image_url",
            "image_url": {
                "url": image_to_data_url(image_path),
            },
        }
    ]
    if prompt:
        content.append(
            {
                "type": "text",
                "text": prompt,
            }
        )

    payload = {
        "model": model,
        "stream": True,
        "messages": [
            {
                "role": "user",
                "content": content,
            }
        ],
    }
    headers = {
        "Accept": "text/event-stream",
        "Content-Type": "application/json",
    }
    if access_code:
        headers["Authorization"] = f"Bearer nk-{access_code}"

    request = urllib.request.Request(
        api_url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            content_type = response.headers.get("Content-Type", "")
            if "text/event-stream" in content_type:
                return read_sse_reply(response)
            return read_json_reply(response.read())
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(read_http_error_message(exc.code, body)) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def read_http_error_message(status_code: int, body: str) -> str:
    if not body:
        return f"HTTP {status_code}"

    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        return f"HTTP {status_code}: {body}"

    message = data.get("message")
    if isinstance(message, str) and message:
        return f"HTTP {status_code}: {message}"
    return f"HTTP {status_code}: {body}"


def format_row_error(exc: Exception) -> str:
    message = str(exc)
    audit_message = extract_iflytek_audit_message(message)
    if audit_message:
        return audit_message
    return f"{type(exc).__name__}: {message}"


def extract_iflytek_audit_message(message: str) -> str:
    if not message:
        return ""

    text = re.sub(r"^HTTP\s+\d+:\s*", "", message.strip(), flags=re.IGNORECASE)
    message_marker = "message="
    if message_marker in text:
        text = text.split(message_marker, 1)[1]
    text = re.sub(r",\s*sid=[^\r\n]*\s*$", "", text, flags=re.DOTALL).strip()
    text = re.sub(
        r"^AuditImageBlockError:\([^)]+\)\s*",
        "",
        text,
        flags=re.DOTALL,
    ).strip()

    if text.startswith("非常抱歉"):
        return text
    return ""


def extract_iflytek_sid(message: str) -> str:
    match = re.search(r"(?:^|[,\s])sid=([^,\r\n]+)", message)
    return match.group(1).strip() if match else ""


def read_sse_reply(response: Any) -> str:
    answer_parts: list[str] = []
    event_lines: list[str] = []

    while True:
        raw_line = response.readline()
        if raw_line == b"":
            break

        line = raw_line.decode("utf-8", errors="replace").rstrip("\r\n")
        if line == "":
            handle_sse_event(event_lines, answer_parts)
            event_lines = []
            continue
        event_lines.append(line)

    if event_lines:
        handle_sse_event(event_lines, answer_parts)

    reply = "".join(answer_parts)
    if not reply:
        raise RuntimeError("Backend returned no reply text.")
    return reply


def handle_sse_event(event_lines: list[str], answer_parts: list[str]) -> None:
    data_lines = []
    for line in event_lines:
        if line.startswith(":"):
            continue
        if line.startswith("data:"):
            data_lines.append(line[5:].lstrip())

    if not data_lines:
        return

    data = "\n".join(data_lines).strip()
    if not data or data == "[DONE]":
        return

    event = json.loads(data)
    if event.get("error"):
        raise RuntimeError(event.get("message") or "Backend returned an error SSE.")

    choices = event.get("choices") or []
    if not choices:
        return

    first = choices[0]
    delta = first.get("delta") or {}
    message = first.get("message") or {}
    content = delta.get("content") or message.get("content") or ""
    if content:
        answer_parts.append(content)


def read_json_reply(body: bytes) -> str:
    text = body.decode("utf-8", errors="replace")
    data = json.loads(text)
    if data.get("error"):
        raise RuntimeError(data.get("message") or text[:500])

    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError("Backend JSON response did not contain choices.")

    first = choices[0]
    message = first.get("message") or {}
    content = message.get("content") or first.get("text") or ""
    if not content:
        raise RuntimeError("Backend JSON response did not contain reply text.")
    return content


def apply_results(
    sheet: Any,
    results: dict[int, Result],
    skipped: list[Result],
    invalid: dict[int, Result],
    reply_column: str,
    columns: ColumnLayout,
) -> None:
    sheet[f"{reply_column}1"] = "回复"
    sheet[f"{columns.status}1"] = "status"
    sheet[f"{columns.latency}1"] = "latency_seconds"
    sheet[f"{columns.error}1"] = "error"
    sheet[f"{columns.sid}1"] = "sid"

    for result in skipped:
        sheet[f"{columns.status}{result.row}"] = result.status
        sheet[f"{columns.error}{result.row}"] = ""
        sheet[f"{columns.sid}{result.row}"] = ""

    for row, result in sorted(invalid.items()):
        sheet[f"{columns.status}{row}"] = result.status
        sheet[f"{columns.error}{row}"] = result.error
        sheet[f"{columns.sid}{row}"] = result.sid

    for row, result in sorted(results.items()):
        if result.status == "success":
            sheet[f"{reply_column}{row}"] = result.reply
        elif result.error.startswith("非常抱歉"):
            sheet[f"{reply_column}{row}"] = result.error
        sheet[f"{columns.status}{row}"] = result.status
        if result.latency_seconds is not None:
            sheet[f"{columns.latency}{row}"] = round(result.latency_seconds, 3)
        sheet[f"{columns.error}{row}"] = result.error
        sheet[f"{columns.sid}{row}"] = result.sid


def print_summary(
    stats: CollectionStats,
    results: dict[int, Result],
    skipped: list[Result],
    invalid: dict[int, Result],
    elapsed_seconds: float,
    output_path: Path,
) -> None:
    all_results = list(results.values()) + list(invalid.values())
    success_count = sum(1 for result in results.values() if result.status == "success")
    failed_count = sum(1 for result in all_results if result.status == "failed")
    skipped_count = len(skipped)
    latencies = [
        result.latency_seconds
        for result in results.values()
        if result.status == "success" and result.latency_seconds is not None
    ]
    avg_latency = sum(latencies) / len(latencies) if latencies else 0.0

    print(f"总行数: {stats.total_rows}")
    print(f"有效图片样本数: {stats.valid_samples}")
    print(f"成功数: {success_count}")
    print(f"失败数: {failed_count}")
    print(f"跳过数: {skipped_count}")
    print(f"图片缺失数: {stats.image_missing_count}")
    print(f"总耗时: {elapsed_seconds:.3f} 秒")
    print(f"平均单题耗时: {avg_latency:.3f} 秒")
    print(f"输出文件路径: {output_path}")


async def async_main() -> int:
    args = parse_args()
    args.image_column = normalize_column(args.image_column, "image-column")
    args.reply_column = normalize_column(args.reply_column, "reply-column")

    if args.concurrency < 1:
        raise ValueError("--concurrency must be at least 1.")
    if args.timeout <= 0:
        raise ValueError("--timeout must be greater than 0.")
    if args.start_row < 2:
        raise ValueError("--start-row must be at least 2.")

    input_path, output_path, image_root = resolve_paths(args)
    workbook = load_workbook(input_path)
    sheet = get_sheet(workbook, args.sheet)
    status_columns = choose_status_columns(sheet, args.reply_column)
    jobs, skipped, invalid, stats = collect_jobs(
        sheet=sheet,
        start_row=args.start_row,
        image_column=args.image_column,
        reply_column=args.reply_column,
        image_root=image_root,
        force=args.force,
    )

    started_at = time.perf_counter()
    results: dict[int, Result] = {}
    try:
        results = await run_jobs(jobs, args)
    finally:
        apply_results(
            sheet=sheet,
            results=results,
            skipped=skipped,
            invalid=invalid,
            reply_column=args.reply_column,
            columns=status_columns,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    print_summary(
        stats=stats,
        results=results,
        skipped=skipped,
        invalid=invalid,
        elapsed_seconds=time.perf_counter() - started_at,
        output_path=output_path,
    )
    has_failed_result = any(
        result.status == "failed" for result in [*results.values(), *invalid.values()]
    )
    return 1 if has_failed_result else 0


def main() -> int:
    try:
        return asyncio.run(async_main())
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001 - CLI should show a concise failure.
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
