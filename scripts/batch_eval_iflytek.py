#!/usr/bin/env python3
"""Batch Excel evaluator for the existing Iflytek imagev4 backend route.

The script intentionally calls the local NextChat backend API instead of
reimplementing Iflytek signing or WebSocket response parsing.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urljoin


DEFAULT_BASE_URL = "http://127.0.0.1:3000"
DEFAULT_API_PATH = "/api/iflytek/v1/chat/completions"
DEFAULT_MODEL = "image@Iflytek"
DEFAULT_QUESTION_COLUMN = "E"
DEFAULT_REPLY_COLUMN = "F"
DEFAULT_STATUS_COLUMN = "G"
DEFAULT_LATENCY_COLUMN = "H"
DEFAULT_ERROR_COLUMN = "I"
DEFAULT_START_ROW = 2
DEFAULT_CONCURRENCY = 5
DEFAULT_TIMEOUT_SECONDS = 120.0
DEFAULT_RETRIES = 2


@dataclass(frozen=True)
class Job:
    row: int
    question: str


@dataclass(frozen=True)
class Result:
    row: int
    status: str
    reply: str = ""
    latency_seconds: float | None = None
    error: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read questions from Excel column E and write independent Iflytek "
            "answers back to column F."
        )
    )
    parser.add_argument("--input", required=True, help="Input .xlsx path.")
    parser.add_argument(
        "--output",
        help=(
            "Output .xlsx path. Defaults to '<input stem>_with_replies.xlsx'. "
            "The input file is never overwritten."
        ),
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to the active worksheet.",
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=DEFAULT_CONCURRENCY,
        help=f"Maximum concurrent requests. Default: {DEFAULT_CONCURRENCY}.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-request rows even when the reply column already has text.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_TIMEOUT_SECONDS,
        help=f"Per-question timeout in seconds. Default: {DEFAULT_TIMEOUT_SECONDS:g}.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_RETRIES,
        help=f"Retries after the first failed attempt. Default: {DEFAULT_RETRIES}.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=DEFAULT_START_ROW,
        help=f"First row containing a question. Default: {DEFAULT_START_ROW}.",
    )
    parser.add_argument(
        "--question-column",
        default=DEFAULT_QUESTION_COLUMN,
        help=f"Question column letter. Default: {DEFAULT_QUESTION_COLUMN}.",
    )
    parser.add_argument(
        "--reply-column",
        default=DEFAULT_REPLY_COLUMN,
        help=f"Reply column letter. Default: {DEFAULT_REPLY_COLUMN}.",
    )
    parser.add_argument(
        "--base-url",
        default=os.getenv("NEXTCHAT_BASE_URL", DEFAULT_BASE_URL),
        help=f"Local NextChat base URL. Default: {DEFAULT_BASE_URL}.",
    )
    parser.add_argument(
        "--api-path",
        default=DEFAULT_API_PATH,
        help=f"Iflytek backend API path. Default: {DEFAULT_API_PATH}.",
    )
    parser.add_argument(
        "--model",
        default=os.getenv("BATCH_IFLYTEK_MODEL", DEFAULT_MODEL),
        help=f"Model sent to the local backend. Default: {DEFAULT_MODEL}.",
    )
    parser.add_argument(
        "--access-code",
        default=os.getenv("NEXTCHAT_ACCESS_CODE", ""),
        help="Optional NextChat access code when CODE is enabled.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not call the backend; write deterministic fake replies.",
    )
    return parser.parse_args()


def normalize_column(value: str, label: str) -> str:
    column = value.strip().upper()
    if not column.isalpha():
        raise ValueError(f"{label} must be an Excel column letter.")
    return column


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_with_replies{input_path.suffix}")


def resolve_paths(args: argparse.Namespace) -> tuple[Path, Path]:
    input_path = Path(args.input).expanduser().resolve()
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else default_output_path(input_path).resolve()
    )

    if input_path == output_path:
        raise ValueError("Output path must be different from input path.")

    return input_path, output_path


def load_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install it with "
            "'python -m pip install openpyxl'."
        ) from exc

    return load_workbook(path)


def get_sheet(workbook: Any, sheet_name: str | None):
    if not sheet_name:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Worksheet '{sheet_name}' not found. Available: {available}")
    return workbook[sheet_name]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def collect_jobs(
    sheet: Any,
    start_row: int,
    question_column: str,
    reply_column: str,
    force: bool,
) -> tuple[list[Job], list[Result], int]:
    jobs: list[Job] = []
    skipped: list[Result] = []
    total_questions = 0

    for row in range(start_row, sheet.max_row + 1):
        question = cell_text(sheet[f"{question_column}{row}"].value)
        if not question:
            continue

        total_questions += 1
        existing_reply = cell_text(sheet[f"{reply_column}{row}"].value)
        if existing_reply and not force:
            skipped.append(Result(row=row, status="skipped", reply=existing_reply))
            continue

        jobs.append(Job(row=row, question=question))

    return jobs, skipped, total_questions


async def run_jobs(
    jobs: list[Job],
    args: argparse.Namespace,
) -> dict[int, Result]:
    semaphore = asyncio.Semaphore(args.concurrency)
    results: dict[int, Result] = {}
    api_url = build_api_url(args.base_url, args.api_path)

    async def run_one(job: Job) -> None:
        async with semaphore:
            if args.dry_run:
                started_at = time.perf_counter()
                await asyncio.sleep(0)
                results[job.row] = Result(
                    row=job.row,
                    status="success",
                    reply=f"[dry-run] {job.question}",
                    latency_seconds=time.perf_counter() - started_at,
                )
                return

            results[job.row] = await request_with_retries(
                job=job,
                api_url=api_url,
                model=args.model,
                access_code=args.access_code,
                timeout_seconds=args.timeout,
                retries=args.retries,
            )

    await asyncio.gather(*(run_one(job) for job in jobs))
    return results


async def request_with_retries(
    job: Job,
    api_url: str,
    model: str,
    access_code: str,
    timeout_seconds: float,
    retries: int,
) -> Result:
    attempts = max(0, retries) + 1
    last_error = ""
    started_at = time.perf_counter()

    for attempt in range(1, attempts + 1):
        try:
            attempt_started_at = time.perf_counter()
            reply = await asyncio.wait_for(
                asyncio.to_thread(
                    call_nextchat_iflytek,
                    api_url,
                    model,
                    job.question,
                    access_code,
                    timeout_seconds,
                ),
                timeout=timeout_seconds,
            )
            return Result(
                row=job.row,
                status="success",
                reply=reply,
                latency_seconds=time.perf_counter() - attempt_started_at,
            )
        except Exception as exc:  # noqa: BLE001 - keep one-row failures isolated.
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt < attempts:
                await asyncio.sleep(min(2 ** (attempt - 1), 5))

    return Result(
        row=job.row,
        status="failed",
        latency_seconds=time.perf_counter() - started_at,
        error=last_error,
    )


def build_api_url(base_url: str, api_path: str) -> str:
    base = base_url.rstrip("/") + "/"
    path = api_path.lstrip("/")
    return urljoin(base, path)


def call_nextchat_iflytek(
    api_url: str,
    model: str,
    question: str,
    access_code: str,
    timeout_seconds: float,
) -> str:
    payload = {
        "model": model,
        "stream": True,
        "messages": [
            {
                "role": "user",
                "content": question,
            }
        ],
    }
    headers = {
        "Accept": "text/event-stream",
        "Content-Type": "application/json",
    }
    if access_code:
        headers["Authorization"] = f"Bearer nk-{access_code}"

    request = urllib.request.Request(
        api_url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            content_type = response.headers.get("Content-Type", "")
            if "text/event-stream" in content_type:
                return read_sse_reply(response)
            return read_json_reply(response.read())
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(read_http_error_message(exc.code, body)) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def read_http_error_message(status_code: int, body: str) -> str:
    if not body:
        return f"HTTP {status_code}"

    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        return f"HTTP {status_code}: {body}"

    message = data.get("message")
    if isinstance(message, str) and message:
        return f"HTTP {status_code}: {message}"
    return f"HTTP {status_code}: {body}"


def read_sse_reply(response: Any) -> str:
    answer_parts: list[str] = []
    event_lines: list[str] = []

    while True:
        raw_line = response.readline()
        if raw_line == b"":
            break

        line = raw_line.decode("utf-8", errors="replace").rstrip("\r\n")
        if line == "":
            handle_sse_event(event_lines, answer_parts)
            event_lines = []
            continue
        event_lines.append(line)

    if event_lines:
        handle_sse_event(event_lines, answer_parts)

    reply = "".join(answer_parts)
    if not reply:
        raise RuntimeError("Backend returned no reply text.")
    return reply


def handle_sse_event(event_lines: list[str], answer_parts: list[str]) -> None:
    data_lines = []
    for line in event_lines:
        if line.startswith(":"):
            continue
        if line.startswith("data:"):
            data_lines.append(line[5:].lstrip())

    if not data_lines:
        return

    data = "\n".join(data_lines).strip()
    if not data or data == "[DONE]":
        return

    event = json.loads(data)
    if event.get("error"):
        raise RuntimeError(event.get("message") or "Backend returned an error SSE.")

    choices = event.get("choices") or []
    if not choices:
        return

    first = choices[0]
    delta = first.get("delta") or {}
    message = first.get("message") or {}
    content = delta.get("content") or message.get("content") or ""
    if content:
        answer_parts.append(content)


def read_json_reply(body: bytes) -> str:
    text = body.decode("utf-8", errors="replace")
    data = json.loads(text)
    if data.get("error"):
        raise RuntimeError(data.get("message") or text[:500])

    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError("Backend JSON response did not contain choices.")

    first = choices[0]
    message = first.get("message") or {}
    content = message.get("content") or first.get("text") or ""
    if not content:
        raise RuntimeError("Backend JSON response did not contain reply text.")
    return content


def apply_results(
    sheet: Any,
    results: dict[int, Result],
    skipped: list[Result],
    reply_column: str,
    status_column: str,
    latency_column: str,
    error_column: str,
) -> None:
    sheet[f"{reply_column}1"] = "回复"
    sheet[f"{status_column}1"] = "status"
    sheet[f"{latency_column}1"] = "latency_seconds"
    sheet[f"{error_column}1"] = "error"

    for result in skipped:
        sheet[f"{status_column}{result.row}"] = result.status
        sheet[f"{error_column}{result.row}"] = ""

    for row, result in sorted(results.items()):
        if result.status == "success":
            sheet[f"{reply_column}{row}"] = result.reply
        sheet[f"{status_column}{row}"] = result.status
        if result.latency_seconds is not None:
            sheet[f"{latency_column}{row}"] = round(result.latency_seconds, 3)
        sheet[f"{error_column}{row}"] = result.error


def print_summary(
    total_questions: int,
    results: dict[int, Result],
    skipped: list[Result],
    elapsed_seconds: float,
    output_path: Path,
) -> None:
    success_count = sum(1 for result in results.values() if result.status == "success")
    failed_count = sum(1 for result in results.values() if result.status == "failed")
    skipped_count = len(skipped)
    latencies = [
        result.latency_seconds
        for result in results.values()
        if result.status == "success" and result.latency_seconds is not None
    ]
    avg_latency = sum(latencies) / len(latencies) if latencies else 0.0

    print(f"总问题数: {total_questions}")
    print(f"成功数: {success_count}")
    print(f"失败数: {failed_count}")
    print(f"跳过数: {skipped_count}")
    print(f"总耗时: {elapsed_seconds:.3f} 秒")
    print(f"平均单题耗时: {avg_latency:.3f} 秒")
    print(f"输出文件路径: {output_path}")


async def async_main() -> int:
    args = parse_args()
    args.question_column = normalize_column(args.question_column, "question-column")
    args.reply_column = normalize_column(args.reply_column, "reply-column")

    if args.concurrency < 1:
        raise ValueError("--concurrency must be at least 1.")
    if args.timeout <= 0:
        raise ValueError("--timeout must be greater than 0.")
    if args.start_row < 2:
        raise ValueError("--start-row must be at least 2.")

    input_path, output_path = resolve_paths(args)
    workbook = load_workbook(input_path)
    sheet = get_sheet(workbook, args.sheet)
    jobs, skipped, total_questions = collect_jobs(
        sheet=sheet,
        start_row=args.start_row,
        question_column=args.question_column,
        reply_column=args.reply_column,
        force=args.force,
    )

    started_at = time.perf_counter()
    results: dict[int, Result] = {}
    try:
        results = await run_jobs(jobs, args)
    finally:
        apply_results(
            sheet=sheet,
            results=results,
            skipped=skipped,
            reply_column=args.reply_column,
            status_column=DEFAULT_STATUS_COLUMN,
            latency_column=DEFAULT_LATENCY_COLUMN,
            error_column=DEFAULT_ERROR_COLUMN,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    print_summary(
        total_questions=total_questions,
        results=results,
        skipped=skipped,
        elapsed_seconds=time.perf_counter() - started_at,
        output_path=output_path,
    )
    return 1 if any(result.status == "failed" for result in results.values()) else 0


def main() -> int:
    try:
        return asyncio.run(async_main())
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001 - CLI should show a concise failure.
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
